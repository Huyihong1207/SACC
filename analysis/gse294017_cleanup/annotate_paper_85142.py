#!/Users/niubi/Desktop/SACC/.venv-scanpy/bin/python
from __future__ import annotations

from collections import Counter
from pathlib import Path

import matplotlib.pyplot as plt
import numpy as np
import pandas as pd
import scanpy as sc
from openpyxl import load_workbook


ROOT = Path("/Users/niubi/Desktop/SACC")
INPUT_H5AD = (
    ROOT
    / "analysis"
    / "gse294017_cleanup"
    / "output"
    / "paper_85142_reclustered"
    / "GSE294017_paper_85142.reclustered.h5ad"
)
FIG4_XLSX = (
    ROOT
    / "paper_assets"
    / "source_data"
    / "Source_Data"
    / "Figure_4_Source_data.xlsx"
)
OUTDIR = (
    ROOT
    / "analysis"
    / "gse294017_cleanup"
    / "output"
    / "paper_85142_annotated"
)

CLUSTER_KEY = "paper_85142_leiden"
UMAP_KEY = "X_umap_paper_85142"
MIN_CLASS_PURITY = 0.80
MIN_SUBTYPE_PURITY = 0.70
MIN_SUBTYPE_CELLS = 20
MIN_MARKER_OVERLAP = 2
TOP_MARKERS_FOR_SUPPORT = 100
MIN_MARKER_LOG2FC = 1.0
MAX_MARKER_P_ADJ = 0.05
MIN_MARKER_PCT_IN = 0.10

CLASS_ORDER = [
    "Malignant_cells",
    "Immune_cells",
    "Alveolar_cells",
    "Fibroblasts/Muscle",
    "Endothelial_cells",
    "Unknown",
]
IMMUNE_ORDER = [
    "Macrophages",
    "T cells",
    "DC",
    "Monocytes",
    "ILC",
    "Plasma cells",
    "pDC",
    "B cells",
    "Mast cells",
]

CLASS_MARKERS = {
    "Alveolar_cells": [
        "EPCAM",
        "KRT8",
        "KRT18",
        "KRT19",
        "MUC1",
        "SCGB1A1",
        "SFTPA1",
        "SFTPA2",
        "SFTPB",
        "SFTPC",
    ],
    "Endothelial_cells": [
        "PECAM1",
        "VWF",
        "EMCN",
        "KDR",
        "ENG",
        "RAMP2",
        "PLVAP",
    ],
    "Fibroblasts/Muscle": [
        "COL1A1",
        "COL1A2",
        "COL3A1",
        "DCN",
        "LUM",
        "COL6A1",
        "ACTA2",
        "TAGLN",
        "MYL9",
        "RGS5",
        "MCAM",
    ],
    "Immune_cells": [
        "PTPRC",
        "LST1",
        "TYROBP",
        "FCER1G",
        "CD3D",
        "CD3E",
        "CD74",
    ],
}
IMMUNE_MARKERS = {
    "Macrophages": [
        "LST1",
        "TYROBP",
        "FCER1G",
        "C1QA",
        "C1QB",
        "C1QC",
        "APOE",
        "CTSD",
    ],
    "T cells": ["CD3D", "CD3E", "TRBC1", "TRBC2", "IL7R", "LTB"],
    "DC": ["FCER1A", "CD1C", "CLEC10A", "HLA-DQA1", "HLA-DQB1"],
    "Monocytes": ["S100A8", "S100A9", "FCN1", "CTSS", "LILRB1"],
    "ILC": ["NKG7", "GNLY", "KLRD1", "XCL1", "XCL2"],
    "Plasma cells": ["MZB1", "JCHAIN", "SDC1", "IGKC", "DERL3"],
    "pDC": ["GZMB", "JCHAIN", "TCF4", "IL3RA", "CLEC4C"],
    "B cells": ["MS4A1", "CD79A", "CD74", "CD37", "CD22", "CD19"],
    "Mast cells": ["TPSAB1", "TPSB2", "KIT", "CPA3", "MS4A2"],
}


def stripped_barcode(obs_name: str) -> str:
    return str(obs_name).rsplit("_", 1)[-1].replace("-1", "")


def natural_cluster_key(value: str) -> tuple[int, str]:
    text = str(value)
    return (int(text), text) if text.isdigit() else (10**9, text)


def load_paper_labels(sheet: str) -> tuple[dict[str, str], Counter]:
    workbook = load_workbook(FIG4_XLSX, data_only=True, read_only=True)
    worksheet = workbook[sheet]
    mapping: dict[str, str] = {}
    counts: Counter = Counter()
    for row_number, row in enumerate(worksheet.iter_rows(values_only=True)):
        if row_number == 0 or not row or row[0] is None:
            continue
        barcode = str(row[0]).strip()
        label = str(row[3]).strip()
        if barcode in mapping:
            raise ValueError(f"Duplicate barcode in {sheet}: {barcode}")
        mapping[barcode] = label
        counts[label] += 1
    return mapping, counts


def excluded_from_display(gene: str) -> bool:
    upper = str(gene).upper()
    return (
        upper == "MALAT1"
        or upper.startswith("MT-")
        or upper.startswith("RPL")
        or upper.startswith("RPS")
    )


def run_find_all_markers(adata: sc.AnnData) -> pd.DataFrame:
    if adata.raw is None:
        raise ValueError("adata.raw is required for full-gene marker analysis")

    sc.tl.rank_genes_groups(
        adata,
        groupby=CLUSTER_KEY,
        groups="all",
        reference="rest",
        method="wilcoxon",
        corr_method="benjamini-hochberg",
        use_raw=True,
        n_genes=adata.raw.n_vars,
        pts=True,
        key_added="rank_genes_groups_raw",
    )
    marker_df = sc.get.rank_genes_groups_df(
        adata,
        group=None,
        key="rank_genes_groups_raw",
    )
    marker_df = marker_df.rename(
        columns={
            "group": "cluster",
            "names": "gene",
            "logfoldchanges": "log2fc",
            "pvals": "p_value",
            "pvals_adj": "p_adj",
            "pct_nz_group": "pct_expr_in",
            "pct_nz_reference": "pct_expr_out",
        }
    )
    marker_df["cluster"] = marker_df["cluster"].astype(str)
    marker_df["display_gene"] = ~marker_df["gene"].map(excluded_from_display)
    return marker_df


def top_gene_sets(filtered_markers: pd.DataFrame) -> dict[str, set[str]]:
    result = {}
    for cluster, frame in filtered_markers.groupby("cluster", observed=True):
        top = frame.sort_values(
            ["log2fc", "pct_expr_in", "p_adj"],
            ascending=[False, False, True],
        ).head(TOP_MARKERS_FOR_SUPPORT)
        result[str(cluster)] = set(top["gene"].astype(str))
    return result


def marker_overlap(
    top_genes: set[str],
    marker_dictionary: dict[str, list[str]],
) -> dict[str, list[str]]:
    return {
        label: sorted(top_genes.intersection(genes))
        for label, genes in marker_dictionary.items()
    }


def majority_summary(series: pd.Series) -> tuple[str, int, float]:
    counts = series.dropna().astype(str).value_counts()
    if counts.empty:
        return "Unmapped", 0, 0.0
    label = str(counts.index[0])
    count = int(counts.iloc[0])
    return label, count, count / int(counts.sum())


def build_cluster_annotations(
    obs: pd.DataFrame,
    filtered_markers: pd.DataFrame,
) -> pd.DataFrame:
    top_genes = top_gene_sets(filtered_markers)
    records = []

    for cluster in sorted(obs[CLUSTER_KEY].astype(str).unique(), key=natural_cluster_key):
        cluster_obs = obs[obs[CLUSTER_KEY].astype(str) == cluster]
        class_label, class_n, class_purity = majority_summary(
            cluster_obs["paper_cell_class"]
        )
        immune_label, immune_n, immune_purity = majority_summary(
            cluster_obs["paper_immune_type"]
        )

        genes = top_genes.get(cluster, set())
        class_overlap = marker_overlap(genes, CLASS_MARKERS)
        immune_overlap = marker_overlap(genes, IMMUNE_MARKERS)
        class_marker_hits = class_overlap.get(class_label, [])
        immune_marker_hits = immune_overlap.get(immune_label, [])

        annotation = "Mixed"
        source = "paper_barcode_reference"
        confidence = "low"
        note = ""

        if class_label == "Unknown" and class_purity >= 0.50:
            annotation = "Unknown"
            confidence = "medium" if class_purity >= MIN_CLASS_PURITY else "low"
        elif class_purity >= MIN_CLASS_PURITY:
            annotation = class_label
            confidence = "high" if class_purity >= 0.90 else "medium"

            if class_label == "Immune_cells":
                enough_reference = immune_n >= MIN_SUBTYPE_CELLS
                subtype_supported = len(immune_marker_hits) >= MIN_MARKER_OVERLAP
                if (
                    enough_reference
                    and immune_purity >= MIN_SUBTYPE_PURITY
                    and subtype_supported
                ):
                    annotation = immune_label
                    source = "paper_Fig4C+raw_markers"
                    confidence = "high" if immune_purity >= 0.85 else "medium"
                elif enough_reference and immune_purity >= MIN_SUBTYPE_PURITY:
                    note = "Paper immune subtype lacks >=2 canonical genes in top markers"
                    confidence = "low"
                else:
                    note = "Insufficient Figure 4C coverage or mixed immune subtype"
                    confidence = "low"
            elif class_label == "Malignant_cells":
                source = "paper_reference_only_inferCNV_CCISM_not_rerun"
            elif class_label in CLASS_MARKERS:
                source = "paper_Fig4B+raw_markers"
                if len(class_marker_hits) < MIN_MARKER_OVERLAP:
                    note = "Paper class lacks >=2 canonical genes in top markers"
                    confidence = "low"
        else:
            note = "No Figure 4B class reaches 80% cluster purity"

        sample_counts = cluster_obs["sample_id"].astype(str).value_counts()
        records.append(
            {
                "cluster": cluster,
                "n_cells": len(cluster_obs),
                "paper_class_majority": class_label,
                "paper_class_majority_n": class_n,
                "paper_class_purity": class_purity,
                "paper_immune_majority": immune_label,
                "paper_immune_majority_n": immune_n,
                "paper_immune_purity": immune_purity,
                "class_marker_hits": ",".join(class_marker_hits),
                "immune_marker_hits": ",".join(immune_marker_hits),
                "cluster_annotation": annotation,
                "annotation_source": source,
                "annotation_confidence": confidence,
                "top_sample": str(sample_counts.index[0]),
                "top_sample_fraction": float(sample_counts.iloc[0] / len(cluster_obs)),
                "annotation_note": note,
            }
        )

    return pd.DataFrame(records)


def write_top_markers_excel(
    filtered_markers: pd.DataFrame,
    annotation_table: pd.DataFrame,
    output_path: Path,
) -> None:
    annotation_lookup = annotation_table.set_index("cluster")[
        "cluster_annotation"
    ].to_dict()
    with pd.ExcelWriter(output_path, engine="openpyxl") as writer:
        annotation_table.to_excel(writer, sheet_name="cluster_annotations", index=False)
        for cluster in sorted(
            filtered_markers["cluster"].unique(),
            key=natural_cluster_key,
        ):
            frame = filtered_markers[
                filtered_markers["cluster"].astype(str) == str(cluster)
            ].copy()
            frame = frame.sort_values(
                ["log2fc", "pct_expr_in", "p_adj"],
                ascending=[False, False, True],
            ).head(50)
            frame.insert(
                1,
                "cluster_annotation",
                annotation_lookup.get(str(cluster), "Unannotated"),
            )
            frame.to_excel(writer, sheet_name=f"cluster_{cluster}"[:31], index=False)


def set_annotation_columns(
    adata: sc.AnnData,
    annotation_table: pd.DataFrame,
) -> None:
    lookup = annotation_table.set_index("cluster")
    cluster_values = adata.obs[CLUSTER_KEY].astype(str)
    for column in [
        "cluster_annotation",
        "annotation_source",
        "annotation_confidence",
    ]:
        adata.obs[column] = cluster_values.map(lookup[column]).astype("category")


def plot_embedding(
    adata: sc.AnnData,
    color: str,
    output_name: str,
    legend_loc: str = "right margin",
) -> None:
    sc.pl.embedding(
        adata,
        basis=UMAP_KEY.removeprefix("X_"),
        color=color,
        legend_loc=legend_loc,
        frameon=False,
        show=False,
    )
    plt.savefig(OUTDIR / output_name, dpi=220, bbox_inches="tight")
    plt.close("all")


def plot_dotplots(adata: sc.AnnData) -> None:
    available = set(adata.raw.var_names.astype(str))
    class_markers = {
        label: [gene for gene in genes if gene in available]
        for label, genes in CLASS_MARKERS.items()
    }
    immune_markers = {
        label: [gene for gene in genes if gene in available]
        for label, genes in IMMUNE_MARKERS.items()
    }

    sc.pl.dotplot(
        adata,
        class_markers,
        groupby="cluster_annotation",
        use_raw=True,
        standard_scale="var",
        show=False,
    )
    plt.savefig(OUTDIR / "dotplot_classic_class_markers.png", dpi=220, bbox_inches="tight")
    plt.close("all")

    immune_mask = adata.obs["cluster_annotation"].isin(IMMUNE_ORDER)
    if int(immune_mask.sum()) > 0:
        sc.pl.dotplot(
            adata[immune_mask],
            immune_markers,
            groupby="cluster_annotation",
            use_raw=True,
            standard_scale="var",
            show=False,
        )
        plt.savefig(
            OUTDIR / "dotplot_classic_immune_markers.png",
            dpi=220,
            bbox_inches="tight",
        )
        plt.close("all")


def build_mapping_audit(
    adata: sc.AnnData,
    paper_class_counts: Counter,
    paper_immune_counts: Counter,
) -> pd.DataFrame:
    barcodes = adata.obs["paper_barcode"].astype(str)
    barcode_counts = barcodes.value_counts()
    observed_class_counts = adata.obs["paper_cell_class"].astype(str).value_counts()
    observed_immune_counts = (
        adata.obs["paper_immune_type"].dropna().astype(str).value_counts()
    )
    records = [
        ("object_cells", adata.n_obs, 85142, adata.n_obs == 85142),
        ("object_clusters", adata.obs[CLUSTER_KEY].nunique(), 26, adata.obs[CLUSTER_KEY].nunique() == 26),
        ("raw_genes_used_for_markers", adata.raw.n_vars, 31396, adata.raw.n_vars == 31396),
        ("object_unique_bare_barcodes", barcodes.nunique(), 85142, barcodes.nunique() == 85142),
        ("object_duplicate_barcode_groups", int((barcode_counts > 1).sum()), 0, int((barcode_counts > 1).sum()) == 0),
        ("object_extra_rows_from_duplicate_barcodes", int((barcode_counts - 1).clip(lower=0).sum()), 0, int((barcode_counts - 1).clip(lower=0).sum()) == 0),
        ("paper_Fig4B_source_rows", sum(paper_class_counts.values()), 85142, sum(paper_class_counts.values()) == 85142),
        ("paper_Fig4C_source_rows", sum(paper_immune_counts.values()), 9205, sum(paper_immune_counts.values()) == 9205),
    ]
    for label in CLASS_ORDER:
        records.append(
            (
                f"Fig4B_{label}_mapped_object_count",
                int(observed_class_counts.get(label, 0)),
                int(paper_class_counts.get(label, 0)),
                int(observed_class_counts.get(label, 0)) == int(paper_class_counts.get(label, 0)),
            )
        )
    for label in IMMUNE_ORDER:
        records.append(
            (
                f"Fig4C_{label}_mapped_object_count",
                int(observed_immune_counts.get(label, 0)),
                int(paper_immune_counts.get(label, 0)),
                int(observed_immune_counts.get(label, 0)) == int(paper_immune_counts.get(label, 0)),
            )
        )
    audit = pd.DataFrame(records, columns=["check", "observed", "expected", "pass"])
    audit["note"] = ""
    failed_mapping = audit["check"].str.contains("mapped_object_count") & ~audit["pass"]
    audit.loc[failed_mapping, "note"] = (
        "Bare 10x barcodes are not sample-qualified in Figure 4 source data; "
        "current object contains cross-sample barcode collisions."
    )
    return audit


def main() -> None:
    OUTDIR.mkdir(parents=True, exist_ok=True)
    adata = sc.read_h5ad(INPUT_H5AD)
    if adata.n_obs != 85142:
        raise ValueError(f"Expected 85142 cells, found {adata.n_obs}")
    if adata.obs[CLUSTER_KEY].nunique() != 26:
        raise ValueError("Expected 26 Leiden clusters")
    if adata.raw is None or adata.raw.n_vars != 31396:
        raise ValueError("Expected adata.raw with 31396 full-gene features")

    class_mapping, class_source_counts = load_paper_labels("Panel_B")
    immune_mapping, immune_source_counts = load_paper_labels("Panel_C")

    adata.obs["paper_barcode"] = [
        stripped_barcode(name) for name in adata.obs_names
    ]
    adata.obs["paper_cell_class"] = (
        adata.obs["paper_barcode"].map(class_mapping).astype("category")
    )
    adata.obs["paper_immune_type"] = (
        adata.obs["paper_barcode"].map(immune_mapping).astype("category")
    )
    adata.obs["paper_label_mapping_status"] = "barcode_only_nonunique_across_samples"

    marker_df = run_find_all_markers(adata)
    filtered_markers = marker_df[
        marker_df["display_gene"]
        & marker_df["log2fc"].gt(MIN_MARKER_LOG2FC)
        & marker_df["p_adj"].lt(MAX_MARKER_P_ADJ)
        & marker_df["pct_expr_in"].ge(MIN_MARKER_PCT_IN)
    ].copy()

    annotation_table = build_cluster_annotations(adata.obs, filtered_markers)
    set_annotation_columns(adata, annotation_table)

    marker_df.to_csv(
        OUTDIR / "findallmarkers_raw_all_genes.csv.gz",
        index=False,
        compression="gzip",
    )
    filtered_markers.to_csv(
        OUTDIR / "findallmarkers_raw_filtered.csv",
        index=False,
    )
    annotation_table.to_csv(
        OUTDIR / "cluster_annotation_summary.tsv",
        sep="\t",
        index=False,
    )
    write_top_markers_excel(
        filtered_markers,
        annotation_table,
        OUTDIR / "top50_markers_by_cluster.xlsx",
    )

    class_cross_tab = pd.crosstab(
        adata.obs[CLUSTER_KEY],
        adata.obs["paper_cell_class"],
    )
    immune_cross_tab = pd.crosstab(
        adata.obs[CLUSTER_KEY],
        adata.obs["paper_immune_type"],
    )
    class_cross_tab.to_csv(OUTDIR / "cluster_by_paper_cell_class.tsv", sep="\t")
    immune_cross_tab.to_csv(OUTDIR / "cluster_by_paper_immune_type.tsv", sep="\t")

    sample_composition = (
        adata.obs.groupby(
            [CLUSTER_KEY, "sample_id"],
            observed=True,
        )
        .size()
        .rename("n_cells")
        .reset_index()
    )
    sample_composition.to_csv(
        OUTDIR / "cluster_sample_composition.tsv",
        sep="\t",
        index=False,
    )

    mapping_audit = build_mapping_audit(
        adata,
        class_source_counts,
        immune_source_counts,
    )
    mapping_audit.to_csv(OUTDIR / "mapping_audit.tsv", sep="\t", index=False)
    pd.DataFrame(
        {
            "parameter": [
                "marker_source",
                "marker_test",
                "marker_reference",
                "min_log2fc",
                "max_p_adj",
                "min_pct_expr_in",
                "class_purity_threshold",
                "immune_subtype_purity_threshold",
            ],
            "value": [
                "adata.raw",
                "wilcoxon",
                "rest",
                MIN_MARKER_LOG2FC,
                MAX_MARKER_P_ADJ,
                MIN_MARKER_PCT_IN,
                MIN_CLASS_PURITY,
                MIN_SUBTYPE_PURITY,
            ],
        }
    ).to_csv(OUTDIR / "annotation_settings.tsv", sep="\t", index=False)

    adata.uns["paper_annotation_limitations"] = {
        "figure4_source_barcode_scope": (
            "Figure 4 source data contains bare 10x barcodes without sample IDs."
        ),
        "current_object_mapping": (
            "Labels are barcode-reference matches and are not exact cell identity "
            "matches where the same barcode occurs in multiple samples."
        ),
        "malignant_assignment": (
            "Malignant labels inherit the paper reference; inferCNV and CCISM were "
            "not rerun in this workflow."
        ),
    }
    adata.write(OUTDIR / "GSE294017_paper_85142.annotated.h5ad")

    plot_embedding(adata, "paper_cell_class", "umap_paper_cell_class.png")
    immune_labeled = adata.obs["paper_immune_type"].notna()
    if int(immune_labeled.sum()) > 0:
        plot_embedding(
            adata[immune_labeled].copy(),
            "paper_immune_type",
            "umap_paper_immune_type.png",
        )
    plot_embedding(
        adata,
        "cluster_annotation",
        "umap_cluster_annotation.png",
        legend_loc="right margin",
    )
    plot_dotplots(adata)

    legacy_marker = (
        ROOT
        / "analysis"
        / "gse294017_cleanup"
        / "output"
        / "paper_85142_reclustered"
        / "markers_paper_85142_leiden.csv"
    )
    if legacy_marker.exists():
        legacy_marker.unlink()

    print(annotation_table.to_string(index=False))
    print("\nMapping audit:")
    print(mapping_audit.to_string(index=False))


if __name__ == "__main__":
    main()
