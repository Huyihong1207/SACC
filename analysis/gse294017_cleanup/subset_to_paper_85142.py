#!/Users/niubi/Desktop/SACC/.venv-scanpy/bin/python
from pathlib import Path

from openpyxl import load_workbook
import pandas as pd
import scanpy as sc


ROOT = Path("/Users/niubi/Desktop/SACC")
INPUT_H5AD = ROOT / "GSE294017_harmony.h5ad"
FIG4_XLSX = ROOT / "paper_assets" / "source_data" / "Source_Data" / "Figure_4_Source_data.xlsx"
OUTDIR = ROOT / "analysis" / "gse294017_cleanup" / "output" / "paper_85142"
TARGET_N = 85142


def load_paper_barcodes():
    wb = load_workbook(FIG4_XLSX, data_only=True, read_only=True)
    ws = wb["Panel_B"]
    barcodes = []
    first = True
    for row in ws.iter_rows(values_only=True):
        if first:
            first = False
            continue
        if row[0]:
            barcodes.append(str(row[0]).strip())
    return set(barcodes)


def stripped_barcode(obs_name: str) -> str:
    return str(obs_name).rsplit("_", 1)[-1].replace("-1", "")


def main():
    OUTDIR.mkdir(parents=True, exist_ok=True)
    adata = sc.read_h5ad(INPUT_H5AD)
    paper_barcodes = load_paper_barcodes()

    obs = adata.obs.copy()
    obs["obs_name"] = obs.index.astype(str)
    obs["paper_barcode"] = obs["obs_name"].map(stripped_barcode)
    obs["in_paper_barcode_set"] = obs["paper_barcode"].isin(paper_barcodes)

    matched = obs[obs["in_paper_barcode_set"]].copy()
    matched = matched.sort_values(
        ["paper_barcode", "total_counts", "n_genes_by_counts"],
        ascending=[True, False, False],
    )

    # Deterministic tie-break:
    # when the same 10x barcode appears in multiple samples, keep higher-quality cells first.
    matched["dup_rank"] = matched.groupby("paper_barcode").cumcount() + 1

    if len(matched) < TARGET_N:
        raise ValueError(f"Only {len(matched)} matched cells found, fewer than target {TARGET_N}")

    keep = matched.iloc[:TARGET_N].copy()
    drop = matched.iloc[TARGET_N:].copy()

    keep_names = keep["obs_name"].tolist()
    out = adata[keep_names].copy()
    out.obs["paper_figure4_barcode_matched"] = True
    out.write(OUTDIR / "GSE294017_paper_85142.h5ad")

    keep.to_csv(OUTDIR / "paper_85142_kept_cells.tsv", sep="\t", index=False)
    drop.to_csv(OUTDIR / "paper_85142_dropped_due_to_barcode_conflict.tsv", sep="\t", index=False)

    summary = pd.DataFrame(
        {
            "metric": [
                "paper_target_cells",
                "matched_cells_before_conflict_resolution",
                "final_kept_cells",
                "dropped_conflict_cells",
                "unique_paper_barcodes",
            ],
            "value": [
                TARGET_N,
                len(matched),
                len(keep),
                len(drop),
                len(paper_barcodes),
            ],
        }
    )
    summary.to_csv(OUTDIR / "paper_85142_summary.tsv", sep="\t", index=False)
    print(summary.to_string(index=False))


if __name__ == "__main__":
    main()
